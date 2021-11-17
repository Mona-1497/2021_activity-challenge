
import openpyxl
from openpyxl import workbook
import mysql.connector

mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password="Mona100%",
    database="ActivityChallengeDB"
)

def read_query(connection, query):
    cursor = connection.cursor()
    cursor.execute(query)
    result = cursor.fetchall()
    return result

mycursor = mydb.cursor(buffered=True)

wb=openpyxl.load_workbook("C:\\Users\\Mona_\\PycharmProjects\\2021_activity-challenge\\View\\aaa.xlsx")
sh=wb['questions']
row=sh.max_row
column=sh.max_column
for i in range(2,row+1):
    for j in range(0,column):
        print(j)

